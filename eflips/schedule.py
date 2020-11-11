# -*- coding: utf-8 -*-
from eflips.settings import global_constants
from eflips.grid import GridSegment, GridPoint, Grid
from collections import deque, OrderedDict
from eflips.misc import list_to_string, hms_str, clear_queue
from eflips import osm
import pandas
import random
import copy
import logging
import collections
import eflips.events
import numpy as np
from math import floor, ceil
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font


def list_segments(schedule):
    table = OrderedDict()
    k = 1
    for trip_node in schedule.root_node.get_children():
        for leg_node in trip_node.get_children():
            leg_pause = leg_node.pause
            num_segments = len(leg_node)
            for index, segment_node in enumerate(leg_node.get_children()):
                origin = segment_node.origin.name
                destination = segment_node.destination.name
                departure = segment_node.departure.getSeconds()
                arrival = segment_node.arrival.getSeconds()
                if index == num_segments - 1:
                    pause = leg_pause
                else:
                    pause = 0
                table.update(
                    {k: {
                        'departure': departure,
                        'arrival': arrival,
                        'origin': origin,
                        'destination': destination,
                        'pause': pause
                    }}
                )
                k += 1
    df_out = pandas.DataFrame.from_dict(table, orient='index')
    return df_out


class ChargingParametersNotPresent(Exception):
    pass


# def select_schedules(schedule_list, criteria, object_type=None):
#     """
#     :param schedule_list: list of schedules to be scanned
#     :param criteria: tuple of three parameters:
#     attribute name, operator, value
#     :param object_type: class type of object to satisfy criteria - optional
#     :return: all of the given schedules, which include one node
#     to satisfy the given criteria
#     """
#     schedule_list_copy = schedule_list.copy()
#     selected_schedule_list = []
#
#     while len(schedule_list_copy) > 0:
#         schedule = schedule_list_copy.pop()
#         if schedule.root_node.satisfies_criteria(
#                 criteria, map_node_name_class_type[object_type]):
#             selected_schedule_list.append(schedule)
#
#     return selected_schedule_list


def sort_schedules_by_departure(schedule_list):
    return sorted(schedule_list, key=lambda schedule: schedule.root_node.departure)


def filter_origin_and_destination_depots(schedule_list):
    consistent_schedule_list = []
    for schedule in schedule_list:
        if schedule.origin_and_destination_depots:
            consistent_schedule_list.append(schedule)

    return consistent_schedule_list


def check_origin_and_destination_depots(schedule_list):
    for schedule in schedule_list:
        if not schedule.origin_and_destination_depots:
            return False

    return True


def cleanup_grid(trip_container, grid_original):
    if isinstance(trip_container, ScheduleContainer):
        raise NotImplementedError('Can\'t handle ScheduleContainers yet, bro')
    elif isinstance(trip_container, TimeTable):
        trip_list = trip_container.get_all()

    grid_new = Grid()
    for trip_node in trip_list:
        # Keep all points and segments contained in the trip list
        for leg_node in trip_node.children:
            for segment_node in leg_node.children:
                origin = segment_node.origin
                destination = segment_node.destination
                segment = segment_node.grid_segment
                if not grid_new.contains_point_id(origin.ID):
                    grid_new.add_point(origin)
                if not grid_new.contains_point_id(destination.ID):
                    grid_new.add_point(destination)
                if not grid_new.contains_segment_id(segment.ID):
                    grid_new.add_segment(segment)

                # Also keep all segments containing a point from the trip list,
                # as well as the connecting points
                # for segment_original in \
                #         grid_original.get_all_segments_by_ID().values():
                #     if not segment.ID == segment_original.ID:
                #         if origin.ID == segment_original.origin.ID or \
                #                 destination.ID == segment_original.origin.ID or \
                #                 origin.ID == segment_original.destination.ID or \
                #                 destination.ID == segment_original.destination.ID:
                #             # Current grid segment shares a location with
                #             # current timetable segment; keep segment and
                #             # its points
                #             if not grid_new.contains_segment_id(
                #                     segment_original.ID):
                #                 grid_new.add_segment(segment_original)
                #             if not grid_new.contains_point_id(
                #                     segment_original.origin.ID):
                #                 grid_new.add_point(segment_original.origin)
                #             if not grid_new.contains_point_id(
                #                     segment_original.destination.ID):
                #                 grid_new.add_point(segment_original.destination)
    for segment_original in \
            grid_original.get_all_segments_by_ID().values():
        for trip_node in trip_list:
            for leg_node in trip_node.children:
                for segment_node in leg_node.children:
                    origin = segment_node.origin
                    destination = segment_node.destination
                    segment = segment_node.grid_segment

                    if not segment.ID == segment_original.ID:
                        if origin.ID == segment_original.origin.ID or \
                                destination.ID == segment_original.origin.ID or \
                                origin.ID == segment_original.destination.ID or \
                                destination.ID == segment_original.destination.ID:
                            # Current grid segment shares a location with
                            # current timetable segment; keep segment and
                            # its points
                            if not grid_new.contains_segment_id(
                                    segment_original.ID):
                                grid_new.add_segment(segment_original)
                            if not grid_new.contains_point_id(
                                    segment_original.origin.ID):
                                grid_new.add_point(segment_original.origin)
                            if not grid_new.contains_point_id(
                                    segment_original.destination.ID):
                                grid_new.add_point(segment_original.destination)
    return grid_new


def trip_info_str(trip_node):
    return '%-10s | %-9s | %-9s | %-5.1f | %-5.1f | %-4.1f | %-4s | %-6s | %-1s | %-32s | %-10s | %-32s | %-10s' % (
        trip_node.ID,
        trip_node.departure.toString_short(),
        trip_node.arrival.toString_short(),
        trip_node.pause/60,
        trip_node.delay/60 if hasattr(trip_node, 'delay') else 0,
        trip_node.distance,
        trip_node.line if hasattr(trip_node, 'line') else 0,
        trip_node.vehicle_type if hasattr(trip_node, 'vehicle_type') else '',
        'E' if trip_node.trip_type == 'empty' else '',
        trip_node.origin.name,
        trip_node.origin.ID,
        trip_node.destination.name, trip_node.destination.ID
    )


def trip_info_str_header():
    header = 'ID         | Dep       | Arr       | Pause | Delay | km   | Line | V.type | E | Origin                           | ID         | Destination                      | ID\n'\
            +'-----------+-----------+-----------+-------+-------+------+------+--------+---+----------------------------------+------------+----------------------------------+----------'
    return header


def schedule_info_str(schedule):
    lines = schedule.root_node.lines
    num_lines = len(lines)
    linestr = ''
    for ind, entry in enumerate(lines):
        linestr += entry
        if ind < num_lines-1:
            linestr += ', '
    return '%-7s | %-9s | %-32s | %-9s | %-32s | %-3d | %-3s | %s' % (
        schedule.root_node.ID,
        schedule.root_node.departure.toString_short(),
        schedule.root_node.origin.name,
        schedule.root_node.arrival.toString_short(),
        schedule.root_node.destination.name,
        schedule.root_node.distance,
        schedule.root_node.vehicle_type,
        linestr
    )


def schedule_info_str_header():
    header = 'ID      | Dep       | Origin                           | Arr       | Destination                      | km  | V.  | Lines\n' +\
             '--------+-----------+----------------------------------+-----------+----------------------------------+-----+-----+-----------------'
    return header


def combine_schedules(list_of_schedules):
    """Put a list of schedules into a ScheduleContainer and create a fresh,
    consistent Grid including all GridPoints and GridSegments present in the
    schedules. Useful for combining schedules from different sources."""
    grid = Grid()
    schedules_out = []
    for schedule in list_of_schedules:
        schedules_out.append(schedule)
        for trip_node in schedule.root_node.children:
            for leg_node in trip_node.children:
                for segment_node in leg_node.children:
                    segment = segment_node.grid_segment
                    origin = segment.origin
                    destination = segment.destination

                    if not grid.contains_point_id(origin.ID):
                        # point is missing in grid completely
                        grid.add_point(origin)
                        origin_new = None
                    elif not grid.contains_point(origin):
                        # point ID exists, but object reference is broken
                        origin_new = grid.get_point(origin.ID)
                        segment.origin = origin_new

                    if not grid.contains_point_id(destination.ID):
                        # point is missing in grid completely
                        grid.add_point(destination)
                        destination_new = None
                    elif not grid.contains_point(destination):
                        # point ID exists, but object reference is broken
                        destination_new = grid.get_point(destination.ID)
                        segment.destination = destination_new

                    if not grid.contains_segment_id(segment.ID):
                        # segment is missing in grid completely
                        grid.add_segment(segment)
                    elif not grid.contains_segment(segment):
                        # segment ID exists, but object reference broken
                        segment_node.grid_segment = grid.get_segment(segment.ID)

                    # if origin_new is None and destination_new is None:
                    #     # object references are clean
                    #     if not grid.contains_segment_id(segment.ID):
                    #         # segment is missing in grid completely
                    #         grid.add_segment(segment)
                    #         rebuild_segment = False
                    #     elif not grid.contains_segment(segment):
                    #         # segment ID exists, but object reference broken
                    #         rebuild_segment = True
                    # else:
                    #     rebuild_segment = True
                    #
                    # if rebuild_segment:
                    #     if origin_new is None:
                    #         origin_new = origin
                    #     if destination_new is None:
                    #         destination_new = destination
                    #     segment_new = GridSegment(segment.ID, origin_new,
                    #                               destination_new)
                    #     grid.add_segment(segment_new)
                    #     segment_node.grid_segment = segment_new
    schedule_container = ScheduleContainer(schedules_out)
    schedule_container.sort_by_departure()

    # re-number schedules:
    schedule_id = 0
    for schedule in schedule_container.get_all():
        schedule.root_node.ID = schedule_id
        schedule_id += 1

    return schedule_container, grid


class ScheduleContainer:
    def __init__(self, schedule_list):
        self._schedule_list = schedule_list
        self._schedules_by_id = dict([(schedule.root_node.ID, schedule) for schedule in schedule_list])

    def __len__(self):
        return len(self._schedule_list)

    @property
    def mean_efficiency(self):
        duration_passenger_trips = 0
        duration_total = 0
        for schedule in self._schedule_list:
            for trip in schedule.root_node.children:
                if trip.trip_type == 'passenger':
                    duration_passenger_trips += trip.duration
                duration_total += trip.duration
        return duration_passenger_trips/duration_total

    @property
    def mean_distance(self):
        distance_total = sum([schedule.root_node.distance
                              for schedule in self._schedule_list])
        return distance_total/len(self._schedule_list)

    @property
    def mean_duration(self):
        duration_total = sum([schedule.root_node.duration
                              for schedule in self._schedule_list])
        return duration_total/len(self._schedule_list)

    def count_trips(self, trip_type=None):
        count = 0
        for schedule in self._schedule_list:
            if trip_type is None:
                count += len(schedule.root_node.children)
            else:
                for trip in schedule.root_node.children:
                    if trip.trip_type == trip_type:
                        count += 1
        return count

    # @property
    def distance(self, trip_type=None):
        if trip_type == None:
            # Return overall distance
            return sum([schedule.root_node.distance for schedule in self._schedule_list])
        else:
            # Return only distance of certain trips (e.g. passenger trips)
            dist = 0
            for schedule in self._schedule_list:
                for trip_node in schedule.root_node.children:
                    if trip_node.trip_type == trip_type:
                        dist += trip_node.distance
            return dist

    @property
    def duration(self):
        return sum([schedule.root_node.duration for schedule in self._schedule_list])

    def generate_timetable(self, trip_type='passenger', zero_pauses=True):
        """Filtering the object ScheduleContainer the keep only one type of trip and make the data construction simpler.

        :param trip_type: type of trip to keep
        :param zero_pauses: decides whether the values of the pauses are set to zero or keep them
        :return: list of trips as an object sorted by departure time
        """
        timetable = TimeTable([])
        for schedule in self._schedule_list:
            schedule_node = copy.copy(schedule).root_node
            for trip_node in schedule_node.get_children():
                if trip_node.trip_type == trip_type:
                    trip_node.vehicle_type = schedule_node.vehicle_type
                    trip_node.parent = None

                    # Set pause to zero
                    if zero_pauses:
                        trip_node.children[-1].pause = 0
                    timetable.add_trip(trip_node)
        timetable.sort_by_departure()
        return timetable

    def get_all(self, copy_objects=False):
        if copy_objects:
            return [Schedule(copy.copy(schedule.root_node)) for schedule in self._schedule_list]
        else:
            return self._schedule_list
        # return self._schedule_list

    def get_schedule(self, ID):
        return self._schedules_by_id[ID]

    def select_schedules(self, criteria, object_type=None):
        """
        Returns all schedules, which include node to satisfy all of the given criteria.
        :param criteria: List of tuples of three parameters: attribute name - string, operator - string, value
        e.g. [('departure', '>', TimeInfo('Tuesday', 3*3600)),
        ('departure', '<=', TimeInfo('Wednesday', 3*3600)),
        ('line', '==', '200')]
        for all operators see Node.satisfies_criterion()
        :param object_type: String of class type of object to satisfy criteria e.g. 'TripNode' - optional
        :return: All schedules, which include one node to satisfy the given criteria
        Jonas Schulte-Mattler
        """
        schedule_list_copy = self._schedule_list.copy()
        selected_schedule_list = []

        while len(schedule_list_copy) > 0:
            schedule = schedule_list_copy.pop()
            if schedule.root_node.satisfies_criteria(
                    criteria, map_node_name_class_type[object_type]):
                selected_schedule_list.append(schedule)

        # self._schedule_list = selected_schedule_list
        return selected_schedule_list

    def sort_by_departure(self):
        self._schedule_list = sorted(self._schedule_list,
                                     key=lambda schedule:
                                     schedule.root_node.departure)

    def remove_zero_legs(self, v_avg=20, v_max=20):
        error_legs = {}
        for schedule in self._schedule_list:
            error_legs[schedule.ID] = schedule.remove_zero_legs(v_avg, v_max)
        return error_legs

    def check_origin_and_destination_depots(self):
        for schedule in self._schedule_list:
            if not schedule.origin_and_destination_depots():
                return False

        return True

    def filter_origin_and_destination_depots(self):
        filtered_schedule_list = []
        for schedule in self._schedule_list:
            if schedule.origin_and_destination_depots():
                filtered_schedule_list.append(schedule)

        self._schedule_list = filtered_schedule_list

    @staticmethod
    def _subtract_pauses_from_legs(leg_nodes, i, v_avg):
        delta_t = leg_nodes[i].distance / (v_avg / 3600)

        n = len(leg_nodes)
        j = 1
        while delta_t > 0 and not (i - j < 0 and i + j > n):
            if i - j > 0:
                if delta_t > leg_nodes[i - j].pause:
                    for k in range(j):
                        leg_nodes[i - k].departure.subSeconds(leg_nodes[i - j].pause)
                    leg_nodes[i].duration_driving += leg_nodes[i - j].pause
                    delta_t -= leg_nodes[i - j].pause
                    leg_nodes[i - j].pause = 0
                else:
                    for k in range(j):
                        leg_nodes[i - k].departure.subSeconds(delta_t)
                    leg_nodes[i].duration_driving += delta_t
                    leg_nodes[i - j].pause -= delta_t
                    delta_t = 0
            if i + j < n:
                if delta_t > leg_nodes[i + j].pause:
                    for k in range(1, j + 1):
                        leg_nodes[i + k].departure.addSeconds(leg_nodes[i + j].pause)
                    leg_nodes[i].duration_driving += leg_nodes[i + j].pause
                    delta_t -= leg_nodes[i + j].pause
                    leg_nodes[i + j].pause = 0
                else:
                    for k in range(1, j + 1):
                        leg_nodes[i + k].departure.addSeconds(delta_t)
                    leg_nodes[i].duration_driving += delta_t
                    leg_nodes[i + j].pause -= delta_t
                    delta_t = 0
            j += 1
        return delta_t

    @staticmethod
    def _shorten_other_legs(leg_nodes, i, delta_t, v_max):
        j = 1
        n = len(leg_nodes)
        while delta_t > 0 and not (i - j < 0 and i + j > n):
            delta_t_leg_max = max(leg_nodes[i - j].duration_driving - (leg_nodes[i - j].distance / (v_max / 3600)), 0)
            if i - j > 0:
                if delta_t > delta_t_leg_max:
                    for k in range(j):
                        leg_nodes[i - k].departure.subSeconds(delta_t_leg_max)
                    leg_nodes[i].duration_driving += delta_t_leg_max
                    delta_t -= delta_t_leg_max
                    leg_nodes[i - j].duration_driving -= delta_t_leg_max
                else:
                    for k in range(j):
                        leg_nodes[i - k].departure.subSeconds(delta_t)
                    leg_nodes[i].duration_driving += delta_t
                    leg_nodes[i - j].duration_driving -= delta_t
                    delta_t = 0
            if i + j < n:
                if delta_t > delta_t_leg_max:
                    for k in range(1, j + 1):
                        leg_nodes[i + k].departure.addSeconds(delta_t_leg_max)
                    leg_nodes[i].duration_driving += delta_t_leg_max
                    delta_t -= delta_t_leg_max
                    leg_nodes[i + j].duration_driving -= delta_t_leg_max
                else:
                    for k in range(1, j + 1):
                        leg_nodes[i + k].departure.addSeconds(delta_t)
                    leg_nodes[i].duration_driving += delta_t
                    leg_nodes[i + j].duration_driving -= delta_t
                    delta_t = 0
            j += 1
        return delta_t

    def remove_zero_legs(self, v_avg=20, v_max=20):
        """
        For all legs having zero duration and nonzero distance,
        the desired duration calculated by v_avg is tried to be set such that
        the pauses of other legs are shortened. If not successful,
        the duration of other legs is set to duration calculated by v_max.
        If that did not work either, leg is added to list being returned.
        :param v_avg: Velocity by which duration of zero legs is calculated
        :param v_max: Velocity by which duration of other legs is calculated
        :return: List of legs with zero duration and nonzero distance for which duration could not be compensated
        Jonas Schulte-Mattler
        """
        logger = logging.getLogger("logger")
        logger.debug("considering average velocity of %d km/h" % v_avg)
        logger.debug("considering maximum velocity of %d km/h" % v_max)

        number_leg_nodes = number_relevant_leg_nodes = number_first_legs = number_last_legs = number_legs_pause = 0
        error_leg_nodes = []

        trip_nodes = self.root_node.children
        first_leg = trip_nodes[0].children[0]
        if first_leg.duration == 0 and first_leg.distance > 0:
            required_time = first_leg.distance / (v_avg / 3600)
            first_leg.duration_driving = required_time
            first_leg.departure.subSeconds(required_time)
            number_first_legs += 1

        last_leg = trip_nodes[-1].children[-1]
        if last_leg.duration == 0 and last_leg.distance > 0:
            required_time = last_leg.distance / (v_avg / 3600)
            last_leg.duration_driving = required_time
            number_last_legs += 1

        for t in range(len(trip_nodes)):
            legs = trip_nodes[t].children
            n = len(legs)
            for i in range(n):
                if legs[i].duration == 0 and legs[i].distance > 0:
                    delta_t = self._subtract_pauses_from_legs(legs, i, v_avg)

                    if delta_t > 0:
                        number_legs_pause += 1
                        delta_t = self._shorten_other_legs(legs, i, delta_t, v_max)

                    if delta_t > 0:
                        error_leg_nodes.append((self, t, i))

                    number_relevant_leg_nodes += 1
            number_leg_nodes += n

        logger.debug("%d of %d legs with duration=0 and length>0 (%.2f%%) found."
                     % (
                     number_relevant_leg_nodes, number_leg_nodes, number_relevant_leg_nodes / number_leg_nodes * 100))
        logger.debug("%d first legs in schedules with duration=0 and length>0 found. "
                     "altered duration departure"
                     % number_first_legs)
        logger.debug("%d last legs in schedules with duration=0 and length>0 found. "
                     "altered driving duration and arrival"
                     % number_last_legs)
        logger.debug(
            "for %d of the %d legs (%.2f%%) the driving duration (calculated by distance and v_avg) "
            "could be compensated for the pauses during the trip" %
            (number_relevant_leg_nodes - number_legs_pause, number_relevant_leg_nodes,
             (number_relevant_leg_nodes - number_legs_pause) / number_relevant_leg_nodes * 100))
        logger.debug("for %d of the %d remaining legs (%.2f%%) the driving duration could be compensated "
                     "for the time difference by driving with v_max during the trip" %
                     (number_legs_pause - len(error_leg_nodes), number_legs_pause,
                      (number_legs_pause - len(error_leg_nodes)) / number_legs_pause * 100))
        logger.warning("for %d legs the driving duration could not be compensated" % (len(error_leg_nodes)))

        return error_leg_nodes

    def export_text(self):
        num_incomplete = 0
        text = ''
        for schedule in self._schedule_list:
            text += schedule.export_text() + '\n\n'
            # if hasattr(schedule.root_node, 'schedule_names'):
            #     title_str = 'Schedule %s (%s)' % (schedule.root_node.ID,
            #         list_to_string(schedule.root_node.schedule_names,
            #                        ', '))
            # else:
            #     title_str = 'Schedule %s' % schedule.root_node.ID
            #
            # title_str += ': %s > %s, %d km' %\
            #              (schedule.root_node.origin.name,
            #               schedule.root_node.destination.name,
            #               schedule.root_node.distance)
            #
            # text += '-----------------------------------------------------------------------------------------------------------------------------------------------------------\n'
            # text += title_str + '\n'
            # text += '-----------------------------------------------------------------------------------------------------------------------------------------------------------\n\n'
            #
            # text += trip_info_str_header() + '\n'
            # for trip_node in schedule.root_node.children:
            #     text += trip_info_str(trip_node) + '\n'
            # text += '\n\n'

            if not (schedule.root_node.origin.type_ == 'depot' and
                    schedule.root_node.destination.type_ == 'depot'):
                num_incomplete += 1

        text += 'Number of incomplete schedules: %d' % num_incomplete
        return text

    def export_text_file(self, file_path):
        with open(file_path, 'w') as file:
            file.write(self.export_text())


class Schedule:
    """
    description...

    note: schedule and vehicle type seem to belong to each other?
    """

    def __init__(self, root_node):
        self.root_node = root_node

    def __copy__(self):
        """This tells the copy.copy() function to include shallow copies of
        the points and segments list. The original GridPoint and GridSegment
        objects in these lists will be retained."""
        new_obj = Schedule(copy.copy(self.root_node))
        return new_obj

    # def satisfies_criteria(self, criteria, object_type):
    #     """
    #     checks if schedule includes object to satisfy given criteria
    #     :param criteria: tuple of three parameters:
    #     attribute name, operator, value
    #     :param object_type: class type of object to satisfy criteria
    #     :return: true if schedule includes object to satisfy given criteria,
    #     false otherwise
    #
    #     Note: similar to dfs, the data structure schedule is not a tree though!
    #     change to bfs by changing stack to queue
    #     """
    #     stack = []
    #     stack.append(self.root_node)
    #     while len(stack) > 0:
    #         node = stack.pop()
    #         if object_type is not None and isinstance(node, object_type):
    #             attribute_value = getattr(node, criteria[0], None)
    #             if attribute_value is not None:
    #                 if criteria[1] == '<' and attribute_value < criteria[2]:
    #                     return True
    #                 if criteria[1] == '<=' and attribute_value <= criteria[2]:
    #                     return True
    #                 if criteria[1] == '>' and attribute_value > criteria[2]:
    #                     return True
    #                 if criteria[1] == '>=' and attribute_value >= criteria[2]:
    #                     return True
    #                 if criteria[1] == '==' and attribute_value == criteria[2]:
    #                     return True
    #         else:
    #             children = node.get_children()
    #             for child in children:
    #                 stack.append(child)

    def check_consistency(self):
        last_leg = self.root_node.children[0].children[0]
        current_leg = last_leg
        while current_leg.get_next_levelorder('right') is not None:
            if isinstance(current_leg.get_next_levelorder('right'), SegmentNode):
                print('Error: SegmentNodes and LegNodes appear at the same tree level!')
            current_leg = current_leg.get_next_levelorder('right')
            departure = last_leg.departure + last_leg.duration
            if not current_leg.departure == departure:
                # +++ DEBUG +++
                try:
                    print('Trip %s: Departures inconsistent! '
                          'Leg %d (%s > %s) departure + duration: %s; '
                          'Leg %d (%s > %s) departure: %s' %
                          (current_leg.parent.ID, last_leg.ID,
                           last_leg.origin.name, last_leg.destination.name,
                           departure.toString_short(),
                           current_leg.ID, current_leg.origin.name,
                           current_leg.destination.name,
                           current_leg.departure.toString_short()))
                except AttributeError:
                    print('Oops')
            last_leg = current_leg

    def origin_and_destination_depots(self):
        return self.root_node.origin.type_ == 'depot' and self.root_node.destination.type_ == 'depot'

    def export_text(self):
        if hasattr(self.root_node, 'schedule_names'):
            title_str = 'Schedule %s (%s)' % (self.root_node.ID,
                                              list_to_string(
                                                  self.root_node.schedule_names,
                                                  ', '))
        else:
            title_str = 'Schedule %s' % self.root_node.ID

        title_str += ': %s > %s, %d km, %s' % \
                     (self.root_node.origin.name,
                      self.root_node.destination.name,
                      self.root_node.distance,
                      self.root_node.vehicle_type)

        text = ''
        text += '-----------------------------------------------------------------------------------------------------------------------------------------------------------\n'
        text += title_str + '\n'
        text += '-----------------------------------------------------------------------------------------------------------------------------------------------------------\n\n'

        text += trip_info_str_header() + '\n'
        for trip_node in self.root_node.children:
            text += trip_info_str(trip_node) + '\n'

        return text


class Node:
    """Universal node class to build trees.

    DJ"""
    def __init__(self, parent, where='right', **kwargs):
        """Create new node. Specify parent=None for root node. Any number of
        attribute-value pairs may be passed through kwargs."""
        self.parent = parent
        self.children = deque()

        for key, value in kwargs.items():
            setattr(self, key, value)

        if parent is not None:
            parent.add_node(self, where=where)

    def __copy__(self):
        """Inheriting classes MUST define their own copy method, otherwise
        copying an object of the respective class will change the type to Node!
        """
        new_obj = Node(None)
        for attr in self.__dict__:
            if not attr == 'children':
                setattr(new_obj, attr, copy.copy(getattr(self, attr)))

        for child in self.children:
            new_obj.add_node(child)

        return new_obj

    #     self._iter_index = 0
    #
    # def __iter__(self):
    #     return iter(self.children)
    #
    # def __next__(self):
    #     if self._iter_index >= len(self.children):
    #         raise StopIteration
    #     res = self.children[self._iter_index]
    #     self._iter_index += 1
    #     return res

    def __len__(self):
        return len(self.children)

    def check_links(self):
        res = True
        for child in self.children:
            if child.parent != self:
                res = False
                break
        return res

    def satisfies_criteria(self, criteria_list, object_type):
        """
        Checks if node or one of its children satisfies all given criteria
        :param criteria_list: list of tuples of three parameters:
        attribute name, operator, value
        e.g. [('departure', '>', TimeInfo('Tuesday', 3*3600)),
        ('departure', '<=', TimeInfo('Wednesday', 3*3600)),
        ('line', '==', '200')]
        for all operators see Node.satisfies_criterion()
        :param object_type: String of class type of object to satisfy criteria e.g. 'TripNode'
        :return: True if node or one of its children satisfies all given criteria, False otherwise

        Note: similar to dfs, the data structure schedule is not a tree though!
        change to bfs by changing stack to queue
        """
        stack = [self]
        while len(stack) > 0:
            node = stack.pop()
            if object_type is not None and isinstance(node, object_type):
                satisfies_all_criteria = True
                for criterium in criteria_list:
                    if not node.satisfies_criterion(criterium):
                        satisfies_all_criteria = False
                        break
                if satisfies_all_criteria:
                    return True
            else:
                for child in node.get_children():
                    stack.append(child)
        return False

    def satisfies_criterion(self, criterion):
        attribute_value = getattr(self, criterion[0], None)
        if attribute_value is not None:
            if criterion[1] == '<' and attribute_value < criterion[2]:
                return True
            if criterion[1] == '<=' and attribute_value <= criterion[2]:
                return True
            if criterion[1] == '>' and attribute_value > criterion[2]:
                return True
            if criterion[1] == '>=' and attribute_value >= criterion[2]:
                return True
            if criterion[1] == '==' and attribute_value == criterion[2]:
                return True
            if criterion[1] == 'in_crit' and attribute_value in criterion[2]:
                return True
            if criterion[1] == 'in_attr' and criterion[2] in attribute_value:
                return True
        return False

    def new_node(self, **kwargs):
        """Generate and add a new child node."""
        new_node = Node(self, **kwargs)
        self.children.append(new_node)
        return new_node

    def add_node(self, node, where='right'):
        """Add a child node already generated elsewhere. Usually there
        should be no need to use this method as it is automatically called
        on the parent when constructing a new node."""
        if where == 'right':
            self.children.append(node)
        elif where == 'left':
            self.children.appendleft(node)
        node.parent = self

    def join(self, node):
        """Join two nodes of the same type, i.e. append the second node's
        children to own children.

        Caution: No sanity checks are performed, the user is responsible for
        not mixing node types!"""
        for child_node in node.children:
            self.add_node(child_node, where='right')

    def remove_node(self, node):
        """Remove a child node."""
        self.children.remove(node)

    def pop_last_node(self):
        return self.children.pop()

    def get_children(self):
        return self.children

    def get_parent(self):
        return self.parent

    def set_parent(self, parent):
        self.parent = parent
        self.parent.children.append(self)

    def siblings(self):
        """Return local 'neighbourhood' of node (i.e., all siblings ÍNCLUDING
        self) and the node's position (index) within the list of siblings."""
        if self.parent is not None:
            siblings = self.parent.children
            pos = siblings.index(self)
            return siblings, pos
        else:
            return self, 0

    def get_next_sibling(self, direction):
        siblings, pos = self.siblings()
        if direction == 'right':
            if pos < len(siblings) - 1:
                return siblings[pos + 1]
            else:
                return None
        elif direction == 'left':
            if pos > 0:
                return siblings[pos - 1]
            else:
                return None
        else:
            logger = logging.getLogger('schedule_logger')
            logger.error("direction must be either 'left' or 'right'")
            raise ValueError("direction must be either 'left' or 'right'")

    def get_right_sibling(self):
        """Return direct sibling to the right (or None if none exists)."""
        siblings, pos = self.siblings()
        if pos < len(siblings)-1:
            return siblings[pos+1]
        else:
            return None

    def get_left_sibling(self):
        """Return direct sibling to the left (or None if none exists)."""
        siblings, pos = self.siblings()
        if pos > 0:
            return siblings[pos-1]
        else:
            return None

    def depth(self):
        """Return the depth of the current node within the tree. The root
        is defined to have a depth of 0."""
        parent = self.parent
        depth = 0
        while parent is not None:
            parent = parent.parent
            depth += 1
        return depth

    def bfs(self):
        """Get all nodes originating from this node through breadth-first
        search (BFS). For testing only.

        Algorithm taken from
        https://www.hackerearth.com/practice/algorithms/graphs/
        breadth-first-search/tutorial/"""
        queue = deque()
        visited = []

        queue.append(self)
        visited.append(self)

        while queue:
            parent_node = queue.popleft()
            for child_node in parent_node.children:
                if child_node not in visited:
                    queue.append(child_node)
                    visited.append(child_node)
                    print(child_node.node_type)

    def get_next_levelorder(self, direction='right'):
        """Return node to the left or right (i.e. next node at current depth),
        even if it descends from another parent node. Useful for getting
        information on the preceding/following segment, leg, trip, ..."""

        # Get current level:
        depth = self.depth()

        if depth == 0:
            return None  # root cannot have any neighbours

        # Check if we have a direct sibling. If not, climb the tree upwards
        # until we find a path leading further to the left/right:
        current_node = self
        next_sibling = current_node.get_next_sibling(direction)
        if next_sibling is None:
            # Find first node upward with a neighbouring sibling
            while not current_node.get_next_sibling(direction):
                current_node = current_node.parent
                # If we arrive at the root, the search will not yield any
                # further results, so return None:
                if current_node.depth() == 0:
                    return None

            # Move one node to the left/right:
            current_node = current_node.get_next_sibling(direction)

            # Now go downward via depth-first search (DFS) and return the first
            # node encountered at the desired depth. DFS algorithm taken from
            # https://www.hackerearth.com/practice/algorithms/graphs/
            # depth-first-search/tutorial/
            queue = deque()
            visited = []

            queue.append(current_node)
            visited.append(current_node)

            while queue:
                # If we are moving right (i.e. we want to go down the leftmost
                # path), we take queue elements from the left; if left, we
                # take them from the right:
                if direction == 'right':
                    current_node = queue.popleft()
                elif direction == 'left':
                    current_node = queue.pop()
                if current_node.depth() == depth:
                    return current_node
                for child_node in current_node.children:
                    if child_node not in visited:
                        queue.append(child_node)
                        visited.append(child_node)

            # If while loop terminates without finding a node of the specified
            # depth, there is no result:
            return None
        else:
            return next_sibling

    def get_left(self):
        """Shortcut for get_next_levelorder(direction='left')."""
        return self.get_next_levelorder(direction='left')

    def get_right(self):
        """Shortcut for get_next_levelorder(direction='right')."""
        return self.get_next_levelorder(direction='right')

    @property
    def origin(self):
        try:
            return self.children[0].origin
        except IndexError:
            return None

    @property
    def destination(self):
        try:
            return self.children[-1].destination
        except IndexError:
            return None

    @property
    def departure(self):
        """Scheduled departure time"""
        try:
            return self.children[0].departure
        except IndexError:
            return None

    @property
    def arrival(self):
        """Scheduled arrival time"""
        try:
            return self.children[-1].arrival
        except IndexError:
            return None

    @property
    def distance(self):
        try:
            return sum([child.distance for child in self.children])
        except AttributeError:
            return 0

    @property
    def duration(self):
        try:
            return sum([child.duration for child in self.children])
        except AttributeError:
            return 0

    @property
    def average_velocity_kmh(self):
        if self.distance == 0:
            return 0
        else:
            return self.distance/self.duration*3600

    # @property
    # def departure(self):
    #     try:
    #         return self.children[0].departure
    #     except IndexError:
    #         return None
    #
    # @property
    # def departure(self):
    #     try:
    #         return self.children[0].departure
    #     except IndexError:
    #         return None


class ScheduleNode(Node):
    def __init__(self, ID, vehicle_type, **kwargs):
        super().__init__(None, ID=ID, vehicle_type=vehicle_type, **kwargs)

    def __copy__(self):
        new_obj = ScheduleNode(self.ID, self.vehicle_type)
        for attr in self.__dict__:
            if attr not in ['ID', 'vehicle_type', 'children']:
                setattr(new_obj, attr, getattr(self, attr))

        for child in self.children:
            new_obj.add_node(copy.copy(child))

        return new_obj

    # def __copy__(self):
    #     """Inheriting classes MUST define their own copy method, otherwise
    #     copying an object of the respective class will change
    #     the type to Node!
    #     """
    #     new_obj = ScheduleNode(self.ID, self.vehicle_type)
    #     if self.parent is not None:
    #         new_obj.parent = self.parent
    #     for child in self.children:
    #         new_obj.children.append(copy.copy(child))

    @property
    def average_velocity_kmh(self):
        """Average speed excluding the pauses after trips."""
        duration = 0
        for trip_node in self.children:
            duration += trip_node.duration - trip_node.pause
        if duration == 0:
            return 0
        else:
            return self.distance/duration*3600

    @property
    def lines(self):
        res = []
        for trip_node in self.children:
            if hasattr(trip_node, 'line'):
                line = trip_node.line
                if not line in res:
                    res.append(line)
        return res


class TripNode(Node):
    def __init__(self, parent, ID, trip_type, **kwargs):
        super().__init__(parent, ID=ID, trip_type=trip_type, **kwargs)

    def __copy__(self):
        new_obj = TripNode(None, self.ID, self.trip_type)
        for attr in self.__dict__:
            if attr not in ['ID', 'trip_type', 'children']:
                setattr(new_obj, attr, getattr(self, attr))

        for child in self.children:
            new_obj.add_node(copy.copy(child))

        return new_obj

    @property
    def average_velocity_kmh(self):
        """Average speed excluding the last pause."""
        if self.duration == 0:
            return 0
        else:
            return self.distance/(self.duration-self.pause)*3600

    @property
    def duration(self):
        """Duration including pause"""
        return sum([leg.duration for leg in self.children])

    @property
    def pause(self):
        """Pause AT END of trip (there may be previous pauses between legs)"""
        return self.children[-1].pause

    @property
    def delay(self):
        # Amount of additional or reduced driving duration; delay can be
        # reduced if there are pauses inbetween legs. However, the pause
        # succeeding the trip (i.e. the last leg) is not included!
        # ^ above is not true, still figuring out how to correctly handle this
        # if len(self.children) > 1:
        #     pause = sum([leg.pause for leg in self.children[0:-2]])
        # else:
        #     pause = 0
        return sum([leg.delay for leg in self.children])


class LegNode(Node):
    def __init__(self, parent, ID, departure_time, pause, **kwargs):
        super().__init__(parent, ID=ID, _departure=departure_time,
                         pause=pause, **kwargs)

    # insert departure_scheduled etc.

    def __copy__(self):
        new_obj = LegNode(None, self.ID, copy.copy(self._departure),
                          self.pause)
        for attr in self.__dict__:
            if attr not in ['ID', '_departure', 'children']:
                setattr(new_obj, attr, getattr(self, attr))

        for child in self.children:
            new_obj.add_node(copy.copy(child))

        return new_obj

    @property
    def average_velocity_kmh(self):
        """Average speed excluding the last pause."""
        if self.duration_driving == 0:
            return 0
        else:
            return self.distance/self.duration_driving*3600

    @property
    def duration_driving(self):
        return sum([segment.duration for segment in self.children])

    @duration_driving.setter
    def duration_driving(self, value):
        segment_duration = value / len(self.children)
        for segment in self.children:
            segment.duration = segment_duration

    @property
    def duration(self):
        return self.duration_driving + self.pause

    @property
    def departure(self):
        return self._departure

    @departure.setter
    def departure(self, value):
        self._departure = value

    @property
    def arrival(self):
        return self.departure + self.duration_driving

    @property
    def delay(self):
        # amount of additional or reduced driving duration; NOT an absolute
        # delay!
        return sum([segment.delay for segment in self.children])


class SegmentNode(Node):
    def __init__(self, parent, grid_segment, duration, **kwargs):
        super().__init__(parent, grid_segment=grid_segment,
                         _duration=duration,
                         **kwargs)

    def __copy__(self):
        new_obj = SegmentNode(None, self.grid_segment, self.duration)
        for attr in self.__dict__:
            if attr not in ['grid_segment', 'duration', 'children']:
                setattr(new_obj, attr, getattr(self, attr))

        for child in self.children:
            new_obj.add_node(copy.copy(child))

        return new_obj

    @property
    def average_velocity_kmh(self):
        """Average speed."""
        if self.duration == 0:
            return 0
        else:
            return self.distance/self.duration*3600

    @property
    def duration(self):
        return self._duration

    @duration.setter
    def duration(self, value):
        self._duration = value

    @property
    def distance(self):
        return self.grid_segment.distance

    @property
    def origin(self):
        return self.grid_segment.origin

    @property
    def destination(self):
        return self.grid_segment.destination

    @property
    def departure(self):
        """Scheduled departure time"""
        if isinstance(self.parent, LegNode):
            leg_dep = self.parent.departure
            dur = 0
            for current_segment in self.parent.get_children():
                # for current_segment in self.parent:
                dep = leg_dep + dur
                dur += current_segment.duration
                if current_segment == self:
                    break
            return dep

            # while current_segment != self:
            #     try:
            #         current_segment = next(self.parent)
            #         dep = leg_dep + dur
            #         dur += current_segment.duration
            #         # arr = leg_dep + dur
            #     except StopIteration:
            #         break
            # return dep
        else:
            return None

    @property
    def arrival(self):
        """Scheduled arrival time"""
        if isinstance(self.parent, LegNode):
            leg_dep = self.parent.departure
            dur = 0
            for current_segment in self.parent.get_children():
                # for current_segment in self.parent:
                dur += current_segment.duration
                arr = leg_dep + dur
                if current_segment == self:
                    break
            return arr
        else:
            return None

    @property
    def delay(self):
        try:
            return self._delay
        except AttributeError:
            return 0

    @delay.setter
    def delay(self, value):
        self._delay = value


# def segment_table(schedule_node):
#     data = OrderedDict()
#     segment.IDx = 0
#     for trip_node in schedule_node:
#         for leg_node in trip_node:
#             pause = leg_node.pause
#             length = len(leg_node)
#             for segment_node, idx in enumerate(leg_node):
#                 data.update(
#                     {segment.IDx:
#                         {'departure_time':
#                         segment_node.departure.getSeconds(),
#                          'arrival_time': segment_node.arrival.getSeconds()}}
#                 )


class Interval:
    """Data container class required for energy flow modeling.
    If the location is the last location, next_location shall be None."""
    def __init__(self, location, duration, next_location, **kwargs):
        self.location = location
        self.duration = duration
        self.next_location = next_location

        for key, value in kwargs.items():
            setattr(self, key, value)


class ChargingScheduleParameterSet:
    """Parameter container used for charging schedule with integrated
    sanity checking."""
    def __init__(self, try_charging, queue_for_charging, charge_full,
                 release_when_full, min_charge_duration):

        # Sanity check:
        if try_charging:
            if charge_full:
                if not queue_for_charging:
                    logger = logging.getLogger('schedule_logger')
                    logger.debug('Warning: queue_for_charging=False conflicts '
                                 'with charge_full=True. Ignoring.')
                    queue_for_charging = True
        else:
            queue_for_charging = False
            charge_full = False
            release_when_full = False

        self.try_charging = try_charging
        self.queue_for_charging = queue_for_charging
        self.charge_full = charge_full
        self.release_when_full = release_when_full
        self.min_charge_duration = min_charge_duration


class ChargingSchedule:
    """Enables location-based charging options."""
    def __init__(self):
        self._data = dict()

    def add(self, schedule_node, location, try_charging,
            queue_for_charging=False,
            charge_full=False,
            release_when_full=False,
            min_charge_duration=0):

        if isinstance(schedule_node, ScheduleNode):
            self._data.update({(schedule_node, location):
                               ChargingScheduleParameterSet(
                                   try_charging, queue_for_charging, charge_full,
                                   release_when_full, min_charge_duration
                               )})
        else:
            raise TypeError('schedule_node must be of type ScheduleNode')

        #
        #
        # # Sanity check:
        # if try_charging:
        #     if charge_full:
        #         if not queue_for_charging:
        #             if global_constants['DEBUG_MSGS']:
        #                 print('Warning: queue_for_charging=False
        #                 conflicts with charge_full=True. Ignoring.')
        #             queue_for_charging = True
        # else:
        #     queue_for_charging = False
        #     charge_full = False
        #     release_when_full = False
        #
        # self._data.update({
        #     (schedule_ID, location):
        #         {'try_charging': try_charging,
        #          'queue_for_charging': queue_for_charging,
        #          'charge_full': charge_full,
        #          'release_when_full': release_when_full}
        # })

    def get(self, schedule_node, location):
        try:
            return self._data[(schedule_node, location)]
        except KeyError:
            # logger = logging.getLogger('schedule_logger')
            # logger.error('ChargingParametersNotPresent')
            raise ChargingParametersNotPresent


class Driver:
    def __init__(self, env, vehicle):
        self.env = env
        self.vehicle = vehicle
        self.vehicle.driver = self
        self.total_time = 0
        self.driving_time = 0
        self.pause_time = 0
        self.additional_paid_time = 0
        self.start_trip_event = eflips.events.EventHandler(self)
        self.end_trip_event = eflips.events.EventHandler(self)
        self.trips_data = {}

    def drive_schedule(self, schedule, additional_paid_time=0):
        logger = logging.getLogger('schedule_logger')
        self.vehicle.mission_list.append(schedule)
        timeout = abs(schedule.root_node.departure) - self.env.now
        yield self.env.timeout(max(timeout, 0))
        logger.debug('t = %d (%s): Driver starting schedule %s' %
                     (self.env.now, hms_str(self.env.now),
                      schedule.root_node.ID))
        if timeout < 0:
            logger.warning('t = %d (%s): Start of schedule %s is delayed!' %
                           (self.env.now, hms_str(self.env.now),
                            schedule.root_node.ID))

        # Switch on
        self.vehicle.ignition_on = True
        self.vehicle.ac_request = True

        # Determine departure time (as TimeInfo)
        wall_time_start = copy.copy(schedule.root_node.departure)
        t0 = self.env.now

        # Go through schedule
        for trip_id, trip_node in enumerate(schedule.root_node.children):
            # log driven trip data for evaluation
            # Caution: Data at the end of the trip is actually logged at the
            # end of the pause following the trip!
            # dt = self.env.now - t0
            trip_data = dict()
            trip_data['trip'] = trip_node
            trip_data['energy_departure'] = \
                self.vehicle.energy_storage_primary.energy.energy
            trip_data['soc_departure'] = \
                self.vehicle.energy_storage_primary.soc
            trip_data['aux_power'] = self.vehicle.aux_power_primary.energy
            trip_data['charging_power'] = \
                self.vehicle.charging_interface_power_primary.energy
            trip_data['departure_time'] = copy.copy(wall_time_start)
            # trip_data['departure_time'].addSeconds(dt)
            trip_data['departure_time_sim'] = self.env.now
            trip_data['delay_departure'] = self.vehicle.delay

            self.start_trip_event(trip_node)
            total_driving_time = 0
            total_break_time = 0
            for leg_node in trip_node.children:
                driving_time, break_time = \
                    yield self.env.process(self.vehicle.drive_leg(leg_node))
                self.driving_time += driving_time
                self.pause_time += break_time
                total_driving_time += driving_time
                total_break_time += break_time
                # yield self.env.process(self._pause(leg_node))
            self.end_trip_event(trip_node)

            trip_data['total_driving_time'] = total_driving_time
            trip_data['total_break_time'] = total_break_time
            # dt = self.env.now - t0
            trip_data['arrival_time'] = copy.copy(wall_time_start)
            # trip_data['arrival_time'].addSeconds(dt-trip_node.pause)
            trip_data['arrival_time_sim'] = self.env.now-trip_node.pause
            trip_data['energy_arrival'] = \
                self.vehicle.energy_storage_primary.energy.energy
            trip_data['soc_arrival'] = self.vehicle.energy_storage_primary.soc
            trip_data['energy_consumed'] = \
                self.vehicle.energy_consumed_primary.energy
            trip_data['delay'] = self.vehicle.delay
            trip_data['delay_in_schedule'] = \
                trip_node.delay if hasattr(trip_node, 'delay') else 0
            trip_data['pause'] = trip_node.pause

            self.trips_data[trip_id] = trip_data

        # Switch off
        # clean events queue before switching off vehicle
        # clear_queue(self.env)
        self.vehicle.ignition_on = False
        self.vehicle.ac_request = False
        Dt = self.env.now - t0
        self.additional_paid_time += additional_paid_time
        self.total_time += Dt + additional_paid_time

    def drive_profile(self, driving_profile, additional_paid_time=0):
        raise NotImplementedError('Driving profile support not yet implemented')


class TimeTable:
    def __init__(self, trip_node_list):
        self._trip_nodes = trip_node_list

    def __len__(self):
        return len(self._trip_nodes)

    def add_trip(self, node):
        if not isinstance(node, TripNode):
            raise TypeError('node must be of type TripNode')
        self._trip_nodes.append(node)

    def sort_by_departure(self):
        self._trip_nodes = sorted(self._trip_nodes,
                                  key=lambda trip_node: trip_node.departure)

    def sort_by_arrival(self):
        self._trip_nodes = sorted(self._trip_nodes,
                                  key=lambda trip_node: trip_node.arrival)

    def get_all(self, copy_objects=False):
        if copy_objects:
            return [copy.copy(trip_node) for trip_node in self._trip_nodes]
        else:
            return self._trip_nodes

    def count_hourly_departures(self, line=None):
        hourly_departures = collections.OrderedDict()
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        for day in days:
            for sec in range(0, 86400, 3600):
                hourly_departures.update({(day, sec): {1: 0, 2: 0}})
        if line is None:
            for trip in self._trip_nodes:
                hourly_departures[(trip.departure.day, trip.departure.hms()[0]*3600)][trip.direction] =\
                    hourly_departures[(trip.departure.day, trip.departure.hms()[0]*3600)][trip.direction]+1
        else:
            for trip in self._trip_nodes:
                if trip.line == line:
                    hourly_departures[(trip.departure.day, trip.departure.hms()[0] * 3600)][trip.direction] = \
                        hourly_departures[(trip.departure.day, trip.departure.hms()[0] * 3600)][trip.direction] + 1
        return hourly_departures

    def select_trips(self, criteria, object_type=None):
        """
        Returns all trips, which include node to satisfy all of the given criteria.
        :param criteria: List of tuples of three parameters: attribute name - string, operator - string, value
        e.g. [('departure', '>', TimeInfo('Tuesday', 3*3600)),
        ('departure', '<=', TimeInfo('Wednesday', 3*3600)),
        ('line', '==', '200')]
        for all operators see Node.satisfies_criterion()
        :param object_type: String of class type of object to satisfy criteria e.g. 'LegNode' - optional
        :return: All trips, which include one node to satisfy the given criteria
        """
        trip_list_copy = self._trip_nodes.copy()
        selected_trip_list = []

        while len(trip_list_copy) > 0:
            trip_node = trip_list_copy.pop()
            if trip_node.satisfies_criteria(
                    criteria, map_node_name_class_type[object_type]):
                selected_trip_list.append(trip_node)

        return selected_trip_list

    def export_text(self):
        text = trip_info_str_header() + '\n'
        for trip_node in self._trip_nodes:
            text += trip_info_str(trip_node) + '\n'
        return text

    def export_text_file(self, file_path):
        with open(file_path, 'w') as file:
            file.write(self.export_text())

    def export_xls_file(self, file_path, file_name, only_terminal_stations=False):
        """This method will export an excel table from the object TimeTable about the trips and the bus stations \
         related to them

        :param file_path: output file's directory
        :param file_name: excel file's name
        :param only_terminal_stations: determine if saving all the stations in the 2nd sheet or only the terminal \
         stations
        :return: excel table contains 2 sheets will be saved in the file_path directory. 1st sheet the information \
         about the trips and the 2nd bus stations information
        """
        ws1_header = ['Fahrt-ID', 'Linie', 'Start', '', '', 'Ziel', '', '', 'Abfahrt', '', 'Ankunft', '',
                          'Länge (km)',
                          'Dauer (min)', 'Pause (min)', 'Fahrzeugtyp', 'Verzögerung (min)']
        ws1_sub_header = [''] * 2 + ['ID', 'Kurzname', 'Name'] * 2 + ['Wochentag', 'Uhrzeit'] * 2 + [''] * 3
        ws1_data_list = [ws1_header, ws1_sub_header]
        ws2_data_set = set()
        ws2_header = ('ID', 'Name', 'Kurzname', 'lon', 'lat')
        for trip_node in self._trip_nodes:
                ws1_data_list.append([trip_node.ID, trip_node.line, trip_node.origin.ID, trip_node.origin.short_name,
                                      trip_node.origin.name, trip_node.destination.ID, trip_node.destination.short_name,
                                      trip_node.destination.name, trip_node.departure.day,
                                      trip_node.departure.toString_hms(),
                                      trip_node.arrival.day, trip_node.arrival.toString_hms(), trip_node.distance,
                                      (trip_node.duration - trip_node.pause) / 60, trip_node.pause / 60,
                                      trip_node.vehicle_type, trip_node.delay / 60])

                if only_terminal_stations:
                    ws2_data_set.add((trip_node.origin.ID, trip_node.origin.name, trip_node.origin.short_name,
                                      trip_node.origin.coords['lon'], trip_node.origin.coords['lat']))
                    ws2_data_set.add((trip_node.destination.ID, trip_node.destination.name, 
                                      trip_node.destination.short_name, trip_node.destination.coords['lon'], 
                                      trip_node.destination.coords['lat']))
                else:
                    for leg in trip_node.get_children():
                        ws2_data_set.add((leg.origin.ID, leg.origin.name, leg.origin.short_name,
                                          leg.origin.coords['lon'], leg.origin.coords['lat']))
                    ws2_data_set.add((leg.destination.ID, leg.destination.name, leg.destination.short_name,
                                      leg.destination.coords['lon'], leg.destination.coords['lat']))
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Fahrplan'
        header_style = Font(size="12", bold=True)
        for row in ws1_data_list:
            ws1.append(row)
        ws1.merge_cells('C1:E1')
        ws1.merge_cells('F1:H1')
        ws1.merge_cells('I1:J1')
        ws1.merge_cells('K1:L1')
        for cell in ws1["1:1"]:
            cell.font = header_style
        for cell in ws1["2:2"]:
            cell.font = header_style
        ws2 = wb.create_sheet(title="Haltestellen")
        ws2.append(ws2_header)
        for row in ws2_data_set:
            ws2.append(row)
        for cell in ws2["1:1"]:
            cell.font = header_style
        if file_name[-5:] != '.xlsx':
            file_name += '.xlsx'
        wb.save(file_path + '\\' + file_name)


def add_random_delays(schedule_list, delay_min, delay_max):
    schedules = copy.copy(schedule_list)
    for schedule in schedules:
        # departure = schedule.root_node.departure
        for trip_node in schedule.root_node.get_children():
            for leg_node in trip_node.get_children():
                delay = random.randint(delay_min, delay_max)
                num_segments = len(leg_node.get_children())
                delay_per_segment = delay/num_segments
                for index, segment_node in enumerate(leg_node.get_children()):
                    # segment_node.duration_scheduled = segment_node.duration
                    # segment_node.duration += delay_per_segment
                    segment_node.delay = delay_per_segment
                # leg_node.departure_scheduled = leg_node.departure
                # leg_node.departure = departure
                # departure = leg_node.arrival + leg_node.pause
    return schedules


map_node_name_class_type = {'ScheduleNode': ScheduleNode,
                            'TripNode': TripNode,
                            'LegNode': LegNode,
                            'SegmentNode': SegmentNode}

class DrivingProfile:
    def __init__(self, name, time, velocity, altitude=None):
        """
        :param name: Name of the profile (string)
        :param time: Time in seconds
        :param velocity: Velocity in m/s
        :param altitude: Altitude in m

        Attributes distance, total_distance and distance_cum are given in m,
        slope in radians,acceleration in m/s²
        """
        if not len(time) == len(velocity):
            raise ValueError('Time and velocity arrays '
                             'must be of the same length')
        if altitude is not None and not len(altitude) == len(time):
            raise ValueError('Time and altitude arrays '
                             'must be of the same length')

        self.name = name
        self.time = np.array(time)
        self.time_delta = np.zeros(len(time))
        self.velocity = np.array(velocity)
        if altitude is None:
            self.altitude = np.zeros(len(time))
        else:
            self.altitude = np.array(altitude)

        # Determine acceleration and slope in radians (backward
        # differentiation) and distance (backward integration)
        self.acceleration = np.zeros(len(time))
        self.slope = np.zeros(len(time))
        self.distance = np.zeros(len(time))
        self.distance_cum = np.zeros(len(time))
        for i in range(1, len(self.time)):
            dv = self.velocity[i] - self.velocity[i-1]
            dt = self.time[i] - self.time[i-1]

            self.time_delta[i] = dt

            # Prevent division by zero:
            if dt == 0:
                self.acceleration[i] = self.acceleration[i-1]
            else:
                self.acceleration[i] = dv/dt

            ds = self.velocity[i]*dt  # m
            self.distance_cum[i] = self.distance_cum[i-1] + ds
            self.distance[i] = ds

            dh = self.altitude[i] - self.altitude[i-1]
            if ds == 0:
                # Vehicle hasn't moved, keep slope (and prevent division
                # by zero):
                self.slope[i] = self.slope[i-1]
            else:
                self.slope[i] = np.arctan(dh/ds)

    def __len__(self):
        return len(self.time)

    @property
    def total_time(self):
        return self.time[-1]

    @property
    def total_distance(self):
        """Return total distance in m"""
        return self.distance_cum[-1]
